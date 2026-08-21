import asyncio
import logging
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.contrib.fsm_storage.memory import MemoryStorage
import psycopg2
from psycopg2.extras import RealDictCursor
import os
import json

# ============ КОНФИГ ============
BOT_TOKEN = "YOUR_BOT_TOKEN_HERE"
ADMIN_IDS = [123456789]
DISPATCHER_IDS = []

DB_CONFIG = {
    "host": "localhost",
    "database": "ustatop_db",
    "user": "postgres",
    "password": "12345"
}

# ============ ЛОГИНГ ============
logging.basicConfig(level=logging.INFO)
storage = MemoryStorage()
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(bot, storage=storage)
dp.middleware.setup(LoggingMiddleware())

# ============ ДАТАБАЗА ============
class Database:
    def __init__(self):
        self.conn = psycopg2.connect(**DB_CONFIG)
        self.cursor = self.conn.cursor()
        self.create_tables()

    def create_tables(self):
        queries = [
            """
            CREATE TABLE IF NOT EXISTS mijozlar (
                id SERIAL PRIMARY KEY,
                tg_id BIGINT UNIQUE,
                ism VARCHAR(100),
                telefon VARCHAR(20),
                manzil TEXT,
                izoh TEXT,
                yaratilgan TIMESTAMP DEFAULT NOW()
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS ustalar (
                id SERIAL PRIMARY KEY,
                tg_id BIGINT UNIQUE,
                ism VARCHAR(100),
                mutaxassislik VARCHAR(100),
                telefon VARCHAR(20),
                reyting FLOAT DEFAULT 0,
                bajarilgan_count INT DEFAULT 0,
                holat BOOLEAN DEFAULT True,
                yaratilgan TIMESTAMP DEFAULT NOW()
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS buyurtmalar (
                id SERIAL PRIMARY KEY,
                mijoz_tg_id BIGINT REFERENCES mijozlar(tg_id),
                usta_id INT REFERENCES ustalar(id),
                dispetcher_id BIGINT,
                matn TEXT,
                manzil TEXT,
                narx DECIMAL(10,2),
                holat VARCHAR(30) DEFAULT 'yangi',
                yaratilgan TIMESTAMP DEFAULT NOW(),
                qabul_qilingan TIMESTAMP,
                boshlangan TIMESTAMP,
                yakunlangan TIMESTAMP,
                bekor_qilingan TIMESTAMP,
                rad_etilgan TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS baholashlar (
                id SERIAL PRIMARY KEY,
                buyurtma_id INT REFERENCES buyurtmalar(id),
                mijoz_tg_id BIGINT REFERENCES mijozlar(tg_id),
                usta_id INT REFERENCES ustalar(id),
                baho INT CHECK (baho >= 1 AND baho <= 5),
                izoh TEXT,
                yaratilgan TIMESTAMP DEFAULT NOW()
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS xabarlar (
                id SERIAL PRIMARY KEY,
                kimdan BIGINT,
                kimgatur VARCHAR(20),
                matn TEXT,
                yuborilgan TIMESTAMP DEFAULT NOW()
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS tolovlar (
                id SERIAL PRIMARY KEY,
                buyurtma_id INT REFERENCES buyurtmalar(id),
                summa DECIMAL(10,2),
                tur VARCHAR(20),
                holat VARCHAR(20) DEFAULT 'kutilmoqda',
                yaratilgan TIMESTAMP DEFAULT NOW()
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS eslatmalar (
                id SERIAL PRIMARY KEY,
                tg_id BIGINT,
                matn TEXT,
                vaqt TIMESTAMP,
                yaratilgan TIMESTAMP DEFAULT NOW()
            )
            """
        ]
        for q in queries:
            self.cursor.execute(q)
        self.conn.commit()

    # ============ МИЖОЗ ============
    def mijoz_qoshish(self, tg_id, ism, telefon, manzil=None, izoh=None):
        self.cursor.execute(
            """INSERT INTO mijozlar (tg_id, ism, telefon, manzil, izoh) 
               VALUES (%s, %s, %s, %s, %s) 
               ON CONFLICT (tg_id) DO UPDATE SET ism=%s, telefon=%s, manzil=%s, izoh=%s""",
            (tg_id, ism, telefon, manzil, izoh, ism, telefon, manzil, izoh)
        )
        self.conn.commit()

    def mijoz_get(self, tg_id):
        self.cursor.execute("SELECT * FROM mijozlar WHERE tg_id=%s", (tg_id,))
        return self.cursor.fetchone()

    def mijoz_all(self):
        self.cursor.execute("SELECT * FROM mijozlar ORDER BY id DESC")
        return self.cursor.fetchall()

    def mijoz_izoh_update(self, tg_id, izoh):
        self.cursor.execute("UPDATE mijozlar SET izoh=%s WHERE tg_id=%s", (izoh, tg_id))
        self.conn.commit()

    # ============ УСТА ============
    def usta_qoshish(self, tg_id, ism, mutaxassislik, telefon):
        self.cursor.execute(
            """INSERT INTO ustalar (tg_id, ism, mutaxassislik, telefon) 
               VALUES (%s, %s, %s, %s) 
               ON CONFLICT (tg_id) DO UPDATE SET ism=%s, mutaxassislik=%s, telefon=%s""",
            (tg_id, ism, mutaxassislik, telefon, ism, mutaxassislik, telefon)
        )
        self.conn.commit()

    def usta_ochirish(self, tg_id):
        self.cursor.execute("DELETE FROM ustalar WHERE tg_id=%s", (tg_id,))
        self.conn.commit()

    def usta_holat_update(self, tg_id, holat):
        self.cursor.execute("UPDATE ustalar SET holat=%s WHERE tg_id=%s", (holat, tg_id))
        self.conn.commit()

    def usta_get(self, tg_id=None, usta_id=None):
        if tg_id:
            self.cursor.execute("SELECT * FROM ustalar WHERE tg_id=%s", (tg_id,))
        elif usta_id:
            self.cursor.execute("SELECT * FROM ustalar WHERE id=%s", (usta_id,))
        else:
            self.cursor.execute("SELECT * FROM ustalar ORDER BY reyting DESC")
            return self.cursor.fetchall()
        return self.cursor.fetchone()

    def usta_barcha(self):
        self.cursor.execute("SELECT * FROM ustalar ORDER BY reyting DESC")
        return self.cursor.fetchall()

    def usta_top(self, mutaxassislik=None):
        if mutaxassislik:
            self.cursor.execute(
                "SELECT * FROM ustalar WHERE mutaxassislik ILIKE %s AND holat=True ORDER BY reyting DESC",
                (f"%{mutaxassislik}%",)
            )
        else:
            self.cursor.execute("SELECT * FROM ustalar WHERE holat=True ORDER BY reyting DESC")
        return self.cursor.fetchall()

    def usta_top_bajargan(self):
        self.cursor.execute("SELECT * FROM ustalar ORDER BY bajarilgan_count DESC LIMIT 10")
        return self.cursor.fetchall()

    # ============ БУЮРТМА ============
    def buyurtma_qoshish(self, mijoz_tg_id, matn, manzil=None, narx=None):
        self.cursor.execute(
            """INSERT INTO buyurtmalar (mijoz_tg_id, matn, manzil, narx) 
               VALUES (%s, %s, %s, %s) RETURNING id""",
            (mijoz_tg_id, matn, manzil, narx)
        )
        self.conn.commit()
        return self.cursor.fetchone()[0]

    def buyurtma_get(self, buyurtma_id):
        self.cursor.execute("SELECT * FROM buyurtmalar WHERE id=%s", (buyurtma_id,))
        return self.cursor.fetchone()

    def buyurtma_update_holat(self, buyurtma_id, holat, usta_id=None, dispetcher_id=None):
        now = datetime.now()
        query = "UPDATE buyurtmalar SET holat=%s"
        params = [holat]
        
        if usta_id:
            query += ", usta_id=%s"
            params.append(usta_id)
        if dispetcher_id:
            query += ", dispetcher_id=%s"
            params.append(dispetcher_id)
        
        holat_map = {
            'qabul_qilingan': ('qabul_qilingan', now),
            'ishda': ('boshlangan', now),
            'yakunlangan': ('yakunlangan', now),
            'bekor_qilingan': ('bekor_qilingan', now),
            'rad_etilgan': ('rad_etilgan', now)
        }
        if holat in holat_map:
            field, time_field = holat_map[holat]
            query += f", {field}=%s"
            params.append(time_field)
        
        query += " WHERE id=%s"
        params.append(buyurtma_id)
        
        self.cursor.execute(query, tuple(params))
        self.conn.commit()
        
        if holat == 'yakunlangan' and usta_id:
            self.cursor.execute(
                "UPDATE ustalar SET bajarilgan_count=bajarilgan_count+1 WHERE id=%s",
                (usta_id,)
            )
            self.conn.commit()

    def buyurtma_usta_change(self, buyurtma_id, yangi_usta_id):
        self.cursor.execute(
            "UPDATE buyurtmalar SET usta_id=%s, holat='qabul_qilingan' WHERE id=%s",
            (yangi_usta_id, buyurtma_id)
        )
        self.conn.commit()

    def buyurtma_all(self, holat=None):
        if holat:
            self.cursor.execute("SELECT * FROM buyurtmalar WHERE holat=%s ORDER BY id DESC", (holat,))
        else:
            self.cursor.execute("SELECT * FROM buyurtmalar ORDER BY id DESC")
        return self.cursor.fetchall()

    def buyurtma_mijoz(self, tg_id):
        self.cursor.execute("SELECT * FROM buyurtmalar WHERE mijoz_tg_id=%s ORDER BY id DESC", (tg_id,))
        return self.cursor.fetchall()

    def buyurtma_usta(self, usta_id):
        self.cursor.execute("SELECT * FROM buyurtmalar WHERE usta_id=%s ORDER BY id DESC", (usta_id,))
        return self.cursor.fetchall()

    def buyurtma_statistika(self):
        self.cursor.execute(
            """SELECT 
                COUNT(*) as jami,
                SUM(CASE WHEN holat='yangi' THEN 1 ELSE 0 END) as yangi,
                SUM(CASE WHEN holat='qabul_qilingan' THEN 1 ELSE 0 END) as qabul,
                SUM(CASE WHEN holat='ishda' THEN 1 ELSE 0 END) as ishda,
                SUM(CASE WHEN holat='yakunlangan' THEN 1 ELSE 0 END) as yakun,
                SUM(CASE WHEN holat='bekor_qilingan' THEN 1 ELSE 0 END) as bekor,
                SUM(CASE WHEN holat='rad_etilgan' THEN 1 ELSE 0 END) as rad
            FROM buyurtmalar"""
        )
        return self.cursor.fetchone()

    def buyurtma_oylik_statistika(self):
        self.cursor.execute(
            """SELECT 
                DATE_TRUNC('month', yaratilgan) as oy,
                COUNT(*) as jami,
                SUM(CASE WHEN holat='yakunlangan' THEN 1 ELSE 0 END) as yakun
            FROM buyurtmalar 
            WHERE yaratilgan >= CURRENT_DATE - INTERVAL '6 months'
            GROUP BY oy
            ORDER BY oy DESC"""
        )
        return self.cursor.fetchall()

    # ============ БАҲОЛАШ ============
    def baho_qoshish(self, buyurtma_id, mijoz_tg_id, usta_id, baho, izoh=None):
        self.cursor.execute(
            "INSERT INTO baholashlar (buyurtma_id, mijoz_tg_id, usta_id, baho, izoh) VALUES (%s, %s, %s, %s, %s)",
            (buyurtma_id, mijoz_tg_id, usta_id, baho, izoh)
        )
        self.conn.commit()
        self.cursor.execute(
            "UPDATE ustalar SET reyting=(SELECT AVG(baho) FROM baholashlar WHERE usta_id=%s) WHERE id=%s",
            (usta_id, usta_id)
        )
        self.conn.commit()

    def baho_get_usta(self, usta_id):
        self.cursor.execute("SELECT AVG(baho) as ortacha, COUNT(*) as soni FROM baholashlar WHERE usta_id=%s", (usta_id,))
        return self.cursor.fetchone()

    # ============ ТЎЛОВ ============
    def tolov_qoshish(self, buyurtma_id, summa, tur):
        self.cursor.execute(
            "INSERT INTO tolovlar (buyurtma_id, summa, tur) VALUES (%s, %s, %s) RETURNING id",
            (buyurtma_id, summa, tur)
        )
        self.conn.commit()
        return self.cursor.fetchone()[0]

    def tolov_holat_update(self, tolov_id, holat):
        self.cursor.execute("UPDATE tolovlar SET holat=%s WHERE id=%s", (holat, tolov_id))
        self.conn.commit()

    def tolov_get_buyurtma(self, buyurtma_id):
        self.cursor.execute("SELECT * FROM tolovlar WHERE buyurtma_id=%s", (buyurtma_id,))
        return self.cursor.fetchall()

    # ============ ЭСЛАТМА ============
    def eslatma_qoshish(self, tg_id, matn, vaqt):
        self.cursor.execute(
            "INSERT INTO eslatmalar (tg_id, matn, vaqt) VALUES (%s, %s, %s) RETURNING id",
            (tg_id, matn, vaqt)
        )
        self.conn.commit()
        return self.cursor.fetchone()[0]

    def eslatma_get_vaqt(self, vaqt):
        self.cursor.execute("SELECT * FROM eslatmalar WHERE vaqt <= %s AND vaqt > %s", 
                           (vaqt, vaqt - timedelta(minutes=1)))
        return self.cursor.fetchall()

    def eslatma_ochirish(self, eslatma_id):
        self.cursor.execute("DELETE FROM eslatmalar WHERE id=%s", (eslatma_id,))
        self.conn.commit()

    # ============ ХАБАР ============
    def xabar_save(self, kimdan, kimgatur, matn):
        self.cursor.execute(
            "INSERT INTO xabarlar (kimdan, kimgatur, matn) VALUES (%s, %s, %s)",
            (kimdan, kimgatur, matn)
        )
        self.conn.commit()

    def close(self):
        self.cursor.close()
        self.conn.close()

db = Database()

# ============ КЛАВИАТУРАЛАР ============
def main_menu():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        KeyboardButton("📝 Буюртма бериш"),
        KeyboardButton("📋 Буюртмаларим"),
        KeyboardButton("👨‍🔧 Усталар"),
        KeyboardButton("⭐ Баҳолаш"),
        KeyboardButton("📞 Диспетчерга мурожаат"),
        KeyboardButton("💬 Отзыв қолдириш"),
        KeyboardButton("👑 Админ панели")
    )
    return kb

def mijoz_menu():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        KeyboardButton("👤 Менинг маълумотларим"),
        KeyboardButton("📝 Буюртма бериш"),
        KeyboardButton("📋 Буюртмаларим"),
        KeyboardButton("⭐ Баҳолаш"),
        KeyboardButton("🔁 Қайта буюртма бериш"),
        KeyboardButton("💬 Отзыв қолдириш"),
        KeyboardButton("📍 Геолокация"),
        KeyboardButton("💳 Онлайн тўлов"),
        KeyboardButton("🔔 Эслатма"),
        KeyboardButton("🏠 Асосий меню")
    )
    return kb

def dispatcher_menu():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        KeyboardButton("📋 Барча буюртмалар"),
        KeyboardButton("🆕 Янги буюртмалар"),
        KeyboardButton("🟡 Қабул қилинганлар"),
        KeyboardButton("🔵 Ишдаги буюртмалар"),
        KeyboardButton("✅ Якунланганлар"),
        KeyboardButton("❌ Бекор қилинганлар"),
        KeyboardButton("🚫 Рад этилганлар"),
        KeyboardButton("📊 Статистика"),
        KeyboardButton("👨‍🔧 Усталар статистикаси"),
        KeyboardButton("📈 Ойлик ҳисобот"),
        KeyboardButton("🏠 Асосий меню")
    )
    return kb

def usta_menu():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        KeyboardButton("👨‍🔧 Менинг профилим"),
        KeyboardButton("📊 Менинг статистикам"),
        KeyboardButton("🏆 Энг кўп буюртма бажарган усталар"),
        KeyboardButton("🟢 Онлайн"),
        KeyboardButton("🔴 Банд"),
        KeyboardButton("📋 Менинг буюртмаларим"),
        KeyboardButton("💰 Даромад"),
        KeyboardButton("🏠 Асосий меню")
    )
    return kb

def admin_menu():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    kb.add(
        KeyboardButton("👨‍🔧 Уста қўшиш"),
        KeyboardButton("👨‍🔧 Уста ўчириш"),
        KeyboardButton("👤 Диспетчер қўшиш"),
        KeyboardButton("👤 Диспетчер ўчириш"),
        KeyboardButton("📊 Умумий статистика"),
        KeyboardButton("📈 Кунлик ҳисобот"),
        KeyboardButton("📈 Ҳафталик ҳисобот"),
        KeyboardButton("📈 Ойлик ҳисобот"),
        KeyboardButton("📢 Барчага хабар"),
        KeyboardButton("📢 Усталарга хабар"),
        KeyboardButton("📢 Мижозларга хабар"),
        KeyboardButton("📥 Excel ҳисобот"),
        KeyboardButton("🏠 Асосий меню")
    )
    return kb

def buyurtma_holat_inline(buyurtma_id):
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🟡 Қабул қилиш", callback_data=f"qabul_{buyurtma_id}"),
        InlineKeyboardButton("🚫 Рад этиш", callback_data=f"rad_{buyurtma_id}"),
        InlineKeyboardButton("🔵 Ишга киришиш", callback_data=f"ishla_{buyurtma_id}"),
        InlineKeyboardButton("✅ Якунлаш", callback_data=f"yakun_{buyurtma_id}"),
        InlineKeyboardButton("❌ Бекор қилиш", callback_data=f"bekor_{buyurtma_id}"),
        InlineKeyboardButton("🔄 Бошқа устага", callback_data=f"change_{buyurtma_id}")
    )
    return kb

def usta_tanlash_inline(ustalar, buyurtma_id):
    kb = InlineKeyboardMarkup(row_width=1)
    for u in ustalar:
        holat = "🟢" if u[6] else "🔴"
        kb.add(InlineKeyboardButton(
            f"{holat} {u[1]} - {u[2]} ⭐{u[4]}",
            callback_data=f"assign_{u[0]}_{buyurtma_id}"
        ))
    return kb

def baholash_inline(buyurtma_id, usta_id):
    kb = InlineKeyboardMarkup(row_width=5)
    for i in range(1, 6):
        kb.insert(InlineKeyboardButton(str(i), callback_data=f"baho_{i}_{buyurtma_id}_{usta_id}"))
    return kb

def tolov_inline(buyurtma_id):
    kb = InlineKeyboardMarkup(row_width=1)
    kb.add(
        InlineKeyboardButton("💳 Тўлов қилиш", callback_data=f"pay_{buyurtma_id}"),
        InlineKeyboardButton("📄 Чекни кўриш", callback_data=f"check_{buyurtma_id}")
    )
    return kb

# ============ ҲОЛАТЛАР ============
from aiogram.dispatcher.filters.state import State, StatesGroup

class BuyurtmaState(StatesGroup):
    ism = State()
    telefon = State()
    manzil = State()
    matn = State()
    narx = State()

class AdminState(StatesGroup):
    usta_qoshish_ism = State()
    usta_qoshish_mutaxassislik = State()
    usta_qoshish_telefon = State()
    usta_qoshish_tg = State()
    dispetcher_qoshish = State()
    xabar_yuborish = State()
    usta_ochirish = State()
    dispetcher_ochirish = State()

class BahoState(StatesGroup):
    buyurtma_id = State()
    izoh = State()

class ChangeUstaState(StatesGroup):
    buyurtma_id = State()
    usta_tanlash = State()

class XabarState(StatesGroup):
    matn = State()

class EslatmaState(StatesGroup):
    matn = State()
    vaqt = State()

class OtzyvState(StatesGroup):
    matn = State()

class GeolocationState(StatesGroup):
    manzil = State()

class TolovState(StatesGroup):
    summa = State()
    tur = State()

# ============ СТАРТ ============
@dp.message_handler(Command("start"))
async def start(message: types.Message):
    user_id = message.from_user.id
    if user_id in ADMIN_IDS:
        await message.answer("👑 *Админ панелига хуш келибсиз!*", reply_markup=admin_menu(), parse_mode="Markdown")
    elif user_id in DISPATCHER_IDS:
        await message.answer("🎧 *Диспетчер панелига хуш келибсиз!*", reply_markup=dispatcher_menu(), parse_mode="Markdown")
    else:
        usta = db.usta_get(tg_id=user_id)
        if usta:
            await message.answer(f"👨‍🔧 *Хуш келибсиз, уста {usta[1]}!*", reply_markup=usta_menu(), parse_mode="Markdown")
        else:
            await message.answer(
                "👤 *Мижоз ботига хуш келибсиз!*\n\n"
                "📝 Буюртма бериш учун тугмани босинг.",
                reply_markup=mijoz_menu(),
                parse_mode="Markdown"
            )

@dp.message_handler(lambda msg: msg.text == "🏠 Асосий меню")
async def back_to_main(message: types.Message):
    await start(message)

# ============ 1. МИЖОЗ: БУЮРТМА БЕРИШ ============
@dp.message_handler(lambda msg: msg.text in ["📝 Буюртма бериш", "🔁 Қайта буюртма бериш"])
async def order_start(message: types.Message, state: FSMContext):
    if message.text == "🔁 Қайта буюртма бериш":
        mijoz = db.mijoz_get(message.from_user.id)
        if mijoz:
            await state.update_data(ism=mijoz[2], telefon=mijoz[3], manzil=mijoz[4] or "")
            await message.answer(f"🔄 *Қайта буюртма*\n\n📝 Буюртма матнини киритинг:", parse_mode="Markdown")
            await BuyurtmaState.matn.set()
            return
    
    await message.answer("👤 *Исмингизни киритинг:*", parse_mode="Markdown")
    await BuyurtmaState.ism.set()

@dp.message_handler(state=BuyurtmaState.ism)
async def get_ism(message: types.Message, state: FSMContext):
    await state.update_data(ism=message.text)
    await message.answer("📱 *Телефон рақамингиз:*", parse_mode="Markdown")
    await BuyurtmaState.telefon.set()

@dp.message_handler(state=BuyurtmaState.telefon)
async def get_telefon(message: types.Message, state: FSMContext):
    await state.update_data(telefon=message.text)
    await message.answer("📍 *Манзилингиз (ёки 'йўқ' ёзинг):*", parse_mode="Markdown")
    await BuyurtmaState.manzil.set()

@dp.message_handler(state=BuyurtmaState.manzil)
async def get_manzil(message: types.Message, state: FSMContext):
    manzil = None if message.text.lower() == 'йўқ' else message.text
    await state.update_data(manzil=manzil)
    await message.answer("📝 *Буюртма матнини киритинг:*", parse_mode="Markdown")
    await BuyurtmaState.matn.set()

@dp.message_handler(state=BuyurtmaState.matn)
async def get_matn(message: types.Message, state: FSMContext):
    await state.update_data(matn=message.text)
    await message.answer("💰 *Нархни киритинг (сўмда, номаълум бўлса 0 ёзинг):*", parse_mode="Markdown")
    await BuyurtmaState.narx.set()

@dp.message_handler(state=BuyurtmaState.narx)
async def get_narx(message: types.Message, state: FSMContext):
    data = await state.get_data()
    try:
        narx = float(message.text)
    except:
        narx = 0
    
    db.mijoz_qoshish(
        tg_id=message.from_user.id,
        ism=data['ism'],
        telefon=data['telefon'],
        manzil=data['manzil']
    )
    
    buyurtma_id = db.buyurtma_qoshish(
        mijoz_tg_id=message.from_user.id,
        matn=data['matn'],
        manzil=data['manzil'],
        narx=narx
    )
    
    holat_map = {
        'yangi': '🆕 Янги',
        'qabul_qilingan': '🟡 Қабул қилинган',
        'ishda': '🔵 Иш жараёнида',
        'yakunlangan': '✅ Якунланган',
        'bekor_qilingan': '❌ Бекор қилинган',
        'rad_etilgan': '🚫 Рад этилган'
    }
    
    txt = (
        f"🆔 *Буюртма №{buyurtma_id}*\n\n"
        f"👤 Мижоз: {data['ism']}\n"
        f"📱 Телефон: {data['telefon']}\n"
        f"📍 Манзил: {data['manzil'] or 'Кўрсатилмаган'}\n"
        f"📝 Матн: {data['matn']}\n"
        f"💰 Нарх: {narx} сўм\n"
        f"📌 Ҳолат: {holat_map['yangi']}\n"
        f"📅 Вақт: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    
    await message.answer(txt, parse_mode="Markdown")
    await state.finish()
    
    # Диспетчерга хабар
    for disp_id in DISPATCHER_IDS:
        try:
            await bot.send_message(
                disp_id,
                f"🔔 *ЯНГИ БУЮРТМА!*\n\n{txt}\n\n⬇️ Уста танлаш:",
                reply_markup=usta_tanlash_inline(db.usta_top(), buyurtma_id),
                parse_mode="Markdown"
            )
        except:
            pass
    
    # Усталарга хабар (онлайнларга)
    for usta in db.usta_top():
        try:
            await bot.send_message(
                usta[1],
                f"🔔 *Янги буюртма!*\n\n🆔 #{buyurtma_id}\n📝 {data['matn'][:50]}...\n💰 {narx} сўм\n\n✅ Қабул қилиш учун /accept_{buyurtma_id}",
                parse_mode="Markdown"
            )
        except:
            pass
    
    await message.answer("✅ *Буюртма қабул қилинди!*")

# ============ 2. МИЖОЗ: БУЮРТМАЛАРИМ ============
@dp.message_handler(lambda msg: msg.text == "📋 Буюртмаларим")
async def my_orders(message: types.Message):
    orders = db.buyurtma_mijoz(message.from_user.id)
    if not orders:
        await message.answer("📭 *Сизда ҳеч қандай буюртма йўқ.*", parse_mode="Markdown")
        return
    
    holat_map = {
        'yangi': '🆕 Янги',
        'qabul_qilingan': '🟡 Қабул қилинган',
        'ishda': '🔵 Иш жараёнида',
        'yakunlangan': '✅ Якунланган',
        'bekor_qilingan': '❌ Бекор қилинган',
        'rad_etilgan': '🚫 Рад этилган'
    }
    
    txt = "📋 *Сизнинг буюртмаларингиз:*\n\n"
    for o in orders[:10]:
        txt += f"🆔 #{o[0]} | {holat_map.get(o[6], o[6])}\n"
        txt += f"📝 {o[4][:40]}...\n\n"
    
    if len(orders) > 10:
        txt += f"\n... ва яна {len(orders)-10} та"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 3. МИЖОЗ: МАЪЛУМОТЛАРИМ ============
@dp.message_handler(lambda msg: msg.text == "👤 Менинг маълумотларим")
async def my_info(message: types.Message):
    mijoz = db.mijoz_get(message.from_user.id)
    if not mijoz:
        await message.answer("⚠️ *Маълумотларингиз топилмади.*", parse_mode="Markdown")
        return
    
    # Устани баҳолаш статистикаси
    orders = db.buyurtma_mijoz(message.from_user.id)
    yakunlangan = len([o for o in orders if o[6] == 'yakunlangan'])
    
    txt = (
        f"👤 *Сизнинг маълумотларингиз:*\n\n"
        f"📛 Исм: {mijoz[2]}\n"
        f"📱 Телефон: {mijoz[3]}\n"
        f"📍 Манзил: {mijoz[4] or 'Кўрсатилмаган'}\n"
        f"📝 Изоҳ: {mijoz[5] or 'Йўқ'}\n"
        f"📋 Жами буюртма: {len(orders)} та\n"
        f"✅ Якунланган: {yakunlangan} та"
    )
    await message.answer(txt, parse_mode="Markdown")

# ============ 4. МИЖОЗ: ИЗОҲ ҚЎШИШ ============
@dp.message_handler(lambda msg: msg.text == "📝 Изоҳ")
async def add_izoh_start(message: types.Message, state: FSMContext):
    await message.answer("📝 *Изоҳ матнини киритинг:*", parse_mode="Markdown")
    await AdminState.usta_qoshish_tg.set()

@dp.message_handler(state=AdminState.usta_qoshish_tg)
async def add_izoh_end(message: types.Message, state: FSMContext):
    db.mijoz_izoh_update(message.from_user.id, message.text)
    await message.answer("✅ *Изоҳ сақланди!*", parse_mode="Markdown")
    await state.finish()

# ============ 5. МИЖОЗ: УСТАЛАР ============
@dp.message_handler(lambda msg: msg.text == "👨‍🔧 Усталар")
async def show_ustalar(message: types.Message):
    ustalar = db.usta_barcha()
    if not ustalar:
        await message.answer("📭 *Усталар мавжуд эмас.*", parse_mode="Markdown")
        return
    
    txt = "👨‍🔧 *Усталар рўйхати:*\n\n"
    for u in ustalar[:10]:
        holat = "🟢 Онлайн" if u[6] else "🔴 Банд"
        txt += f"📛 {u[1]}\n🔧 {u[2]}\n⭐ {u[4]} | 🏆 {u[5]} та\n📌 {holat}\n\n"
    
    if len(ustalar) > 10:
        txt += f"\n... ва яна {len(ustalar)-10} та уста"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 6. МИЖОЗ: БАҲОЛАШ ============
@dp.message_handler(lambda msg: msg.text == "⭐ Баҳолаш")
async def baholash_start(message: types.Message, state: FSMContext):
    await message.answer("🆔 *Баҳолаш учун буюртма рақамини киритинг:*", parse_mode="Markdown")
    await BahoState.buyurtma_id.set()

@dp.message_handler(state=BahoState.buyurtma_id)
async def baholash_get_order(message: types.Message, state: FSMContext):
    try:
        buyurtma_id = int(message.text)
        order = db.buyurtma_get(buyurtma_id)
        if not order or order[1] != message.from_user.id or order[6] != 'yakunlangan':
            await message.answer("⚠️ *Буюртма топилмади ёки ҳали якунланмаган.*", parse_mode="Markdown")
            await state.finish()
            return
        await state.update_data(buyurtma_id=buyurtma_id, usta_id=order[2])
        await message.answer(
            "⭐ *Баҳо беринг (1-5):*",
            reply_markup=baholash_inline(buyurtma_id, order[2]),
            parse_mode="Markdown"
        )
        await BahoState.izoh.set()
    except:
        await message.answer("⚠️ *Буюртма рақамини киритинг!*", parse_mode="Markdown")

@dp.callback_query_handler(lambda c: c.data.startswith("baho_"), state=BahoState.izoh)
async def baholash_set(callback: types.CallbackQuery, state: FSMContext):
    _, baho, buyurtma_id, usta_id = callback.data.split("_")
    baho = int(baho)
    buyurtma_id = int(buyurtma_id)
    usta_id = int(usta_id)
    
    await callback.message.answer("💬 *Изоҳ қолдиринг (ёки 'йўқ' ёзинг):*", parse_mode="Markdown")
    
    @dp.message_handler(state=BahoState.izoh)
    async def baholash_izoh(message: types.Message, state: FSMContext):
        izoh = message.text if message.text.lower() != 'йўқ' else None
        db.baho_qoshish(buyurtma_id, callback.from_user.id, usta_id, baho, izoh)
        await message.answer(f"✅ *Рахмат! Сиз {baho} ★ бердингиз.*", parse_mode="Markdown")
        await state.finish()

# ============ 7. МИЖОЗ: ОТЗЫВ (БАҲО ВА ИЗОҲ) ============
@dp.message_handler(lambda msg: msg.text == "💬 Отзыв қолдириш")
async def otzyv_start(message: types.Message, state: FSMContext):
    await message.answer("🆔 *Отзыв қолдириш учун буюртма рақамини киритинг:*", parse_mode="Markdown")
    await BahoState.buyurtma_id.set()

# ============ 8. МИЖОЗ: ГЕОЛОКАЦИЯ ============
@dp.message_handler(lambda msg: msg.text == "📍 Геолокация")
async def geolocation_start(message: types.Message):
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton("📍 Жойлашувни юбориш", request_location=True))
    await message.answer("📍 *Жойлашувингизни юборинг ёки манзилни ёзинг:*", reply_markup=kb, parse_mode="Markdown")

@dp.message_handler(content_types=['location'])
async def geolocation_get(message: types.Message):
    location = message.location
    await message.answer(
        f"📍 *Жойлашув қабул қилинди!*\n"
        f"🌐 Latitude: {location.latitude}\n"
        f"🌐 Longitude: {location.longitude}\n\n"
        f"🗺 Google Maps: https://maps.google.com/?q={location.latitude},{location.longitude}",
        parse_mode="Markdown"
    )

# ============ 9. МИЖОЗ: ОНЛАЙН ТЎЛОВ ============
@dp.message_handler(lambda msg: msg.text == "💳 Онлайн тўлов")
async def online_tolov_start(message: types.Message):
    orders = db.buyurtma_mijoz(message.from_user.id)
    yakunlangan = [o for o in orders if o[6] == 'yakunlangan']
    
    if not yakunlangan:
        await message.answer("📭 *Тўлов қилиш учун якунланган буюртмалар йўқ.*", parse_mode="Markdown")
        return
    
    kb = InlineKeyboardMarkup(row_width=1)
    for o in yakunlangan:
        kb.add(InlineKeyboardButton(
            f"💳 #{o[0]} - {o[5] or 0} сўм",
            callback_data=f"pay_{o[0]}"
        ))
    
    await message.answer("💳 *Тўлов қилиш учун буюртмани танланг:*", reply_markup=kb, parse_mode="Markdown")

@dp.callback_query_handler(lambda c: c.data.startswith("pay_"))
async def online_tolov_pay(callback: types.CallbackQuery):
    buyurtma_id = int(callback.data.split("_")[1])
    order = db.buyurtma_get(buyurtma_id)
    
    if not order:
        await callback.message.edit_text("⚠️ *Буюртма топилмади.*", parse_mode="Markdown")
        return
    
    summa = float(order[5] or 0)
    if summa == 0:
        await callback.message.edit_text("⚠️ *Бу буюртма бепул!*", parse_mode="Markdown")
        return
    
    # Тўлов жараёни
    tolov_id = db.tolov_qoshish(buyurtma_id, summa, 'online')
    
    await callback.message.edit_text(
        f"💳 *Тўлов маълумотлари:*\n\n"
        f"🆔 Буюртма: #{buyurtma_id}\n"
        f"💰 Сумма: {summa} сўм\n"
        f"📌 Ҳолат: КУТИЛМОҚДА\n\n"
        f"Тўловни тасдиқлаш учун /confirm_{tolov_id}",
        parse_mode="Markdown"
    )

@dp.message_handler(lambda msg: msg.text.startswith("/confirm_"))
async def confirm_tolov(message: types.Message):
    tolov_id = int(message.text.split("_")[1])
    db.tolov_holat_update(tolov_id, 'тўланди')
    await message.answer("✅ *Тўлов тасдиқланди! Рахмат.*", parse_mode="Markdown")

# ============ 10. МИЖОЗ: ЭСЛАТМА ============
@dp.message_handler(lambda msg: msg.text == "🔔 Эслатма")
async def eslatma_start(message: types.Message, state: FSMContext):
    await message.answer("📝 *Эслатма матнини киритинг:*", parse_mode="Markdown")
    await EslatmaState.matn.set()

@dp.message_handler(state=EslatmaState.matn)
async def eslatma_matn(message: types.Message, state: FSMContext):
    await state.update_data(matn=message.text)
    await message.answer("⏰ *Вақтни киритинг (DD.MM.YYYY HH:MM):*", parse_mode="Markdown")
    await EslatmaState.vaqt.set()

@dp.message_handler(state=EslatmaState.vaqt)
async def eslatma_vaqt(message: types.Message, state: FSMContext):
    try:
        vaqt = datetime.strptime(message.text, "%d.%m.%Y %H:%M")
        data = await state.get_data()
        db.eslatma_qoshish(message.from_user.id, data['matn'], vaqt)
        await message.answer(f"✅ *Эслатма сақланди!*\n\n📝 {data['matn']}\n⏰ {vaqt.strftime('%d.%m.%Y %H:%M')}", parse_mode="Markdown")
        await state.finish()
    except:
        await message.answer("⚠️ *Вақт формати хато! DD.MM.YYYY HH:MM*", parse_mode="Markdown")

# ============ 11. МИЖОЗ: ДИСПЕТЧЕРГА МУРОЖААТ ============
@dp.message_handler(lambda msg: msg.text == "📞 Диспетчерга мурожаат")
async def contact_dispatcher(message: types.Message):
    if DISPATCHER_IDS:
        txt = "📞 *Диспетчер билан боғланиш:*\n\n"
        for d in DISPATCHER_IDS:
            try:
                user = await bot.get_chat(d)
                txt += f"👤 {user.full_name}\n"
            except:
                txt += f"👤 ID: {d}\n"
        await message.answer(txt, parse_mode="Markdown")
    else:
        await message.answer("📞 *Диспетчер ҳозирча мавжуд эмас. Кейинроқ уриниб кўринг.*", parse_mode="Markdown")

# ============ 12. УСТА: ПРОФИЛЬ ============
@dp.message_handler(lambda msg: msg.text == "👨‍🔧 Менинг профилим")
async def usta_profile(message: types.Message):
    usta = db.usta_get(tg_id=message.from_user.id)
    if not usta:
        await message.answer("⚠️ *Сиз уста эмассиз.*", parse_mode="Markdown")
        return
    
    holat = "🟢 Онлайн" if usta[6] else "🔴 Банд"
    baho = db.baho_get_usta(usta[0])
    
    txt = (
        f"👨‍🔧 *Уста профили*\n\n"
        f"📛 Исм: {usta[1]}\n"
        f"🔧 Ихтисос: {usta[2]}\n"
        f"📱 Телефон: {usta[3]}\n"
        f"⭐ Рейтинг: {usta[4]:.1f}\n"
        f"📊 Баҳолар: {baho[1] or 0} та\n"
        f"🏆 Бажарган: {usta[5]} та\n"
        f"📌 Ҳолат: {holat}"
    )
    await message.answer(txt, parse_mode="Markdown")

# ============ 13. УСТА: СТАТИСТИКА ============
@dp.message_handler(lambda msg: msg.text == "📊 Менинг статистикам")
async def usta_statistika(message: types.Message):
    usta = db.usta_get(tg_id=message.from_user.id)
    if not usta:
        await message.answer("⚠️ *Сиз уста эмассиз.*", parse_mode="Markdown")
        return
    
    orders = db.buyurtma_usta(usta[0])
    holat_count = {}
    for o in orders:
        holat_count[o[6]] = holat_count.get(o[6], 0) + 1
    
    # Даромад
    daromad = 0
    for o in orders:
        if o[6] == 'yakunlangan' and o[5]:
            daromad += float(o[5])
    
    txt = (
        f"📊 *Уста статистикаси*\n\n"
        f"🏆 Бажарган: {usta[5]} та\n"
        f"⭐ Рейтинг: {usta[4]:.1f}\n"
        f"💰 Даромад: {daromad:,.0f} сўм\n"
        f"📋 Жами буюртма: {len(orders)} та\n"
        f"🆕 Янги: {holat_count.get('yangi', 0)} та\n"
        f"🟡 Қабул: {holat_count.get('qabul_qilingan', 0)} та\n"
        f"🔵 Ишда: {holat_count.get('ishda', 0)} та\n"
        f"✅ Якун: {holat_count.get('yakunlangan', 0)} та"
    )
    await message.answer(txt, parse_mode="Markdown")

# ============ 14. УСТА: ЭНГ КЎП БАЖАРГАНЛАР ============
@dp.message_handler(lambda msg: msg.text == "🏆 Энг кўп буюртма бажарган усталар")
async def usta_top_bajargan(message: types.Message):
    ustalar = db.usta_top_bajargan()
    if not ustalar:
        await message.answer("📭 *Усталар мавжуд эмас.*", parse_mode="Markdown")
        return
    
    txt = "🏆 *Энг кўп буюртма бажарган усталар:*\n\n"
    for i, u in enumerate(ustalar, 1):
        holat = "🟢" if u[6] else "🔴"
        txt += f"{i}. {holat} {u[1]}\n🔧 {u[2]} | ⭐{u[4]:.1f} | 🏆{u[5]} та\n\n"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 15. УСТА: ДАРОМАД ============
@dp.message_handler(lambda msg: msg.text == "💰 Даромад")
async def usta_daromad(message: types.Message):
    usta = db.usta_get(tg_id=message.from_user.id)
    if not usta:
        await message.answer("⚠️ *Сиз уста эмассиз.*", parse_mode="Markdown")
        return
    
    orders = db.buyurtma_usta(usta[0])
    daromad = 0
    yakunlangan = 0
    
    for o in orders:
        if o[6] == 'yakunlangan':
            yakunlangan += 1
            if o[5]:
                daromad += float(o[5])
    
    txt = (
        f"💰 *Даромад ҳисоботи*\n\n"
        f"👨‍🔧 Уста: {usta[1]}\n"
        f"🏆 Бажарган: {yakunlangan} та\n"
        f"💰 Жами даромад: {daromad:,.0f} сўм\n"
        f"📊 Ўртача нарх: {daromad/yakunlangan:,.0f} сўм" if yakunlangan > 0 else "📊 Ўртача нарх: 0 сўм"
    )
    await message.answer(txt, parse_mode="Markdown")

# ============ 16. УСТА: ОНЛАЙН/БАНД ============
@dp.message_handler(lambda msg: msg.text in ["🟢 Онлайн", "🔴 Банд"])
async def usta_holat_update(message: types.Message):
    usta = db.usta_get(tg_id=message.from_user.id)
    if not usta:
        await message.answer("⚠️ *Сиз уста эмассиз.*", parse_mode="Markdown")
        return
    
    holat = True if message.text == "🟢 Онлайн" else False
    db.usta_holat_update(message.from_user.id, holat)
    await message.answer(f"✅ *Ҳолат янгиланди: {message.text}*", parse_mode="Markdown")

# ============ 17. УСТА: МЕНИНГ БУЮРТМАЛАРИМ ============
@dp.message_handler(lambda msg: msg.text == "📋 Менинг буюртмаларим")
async def usta_orders(message: types.Message):
    usta = db.usta_get(tg_id=message.from_user.id)
    if not usta:
        await message.answer("⚠️ *Сиз уста эмассиз.*", parse_mode="Markdown")
        return
    
    orders = db.buyurtma_usta(usta[0])
    if not orders:
        await message.answer("📭 *Сизга ҳеч қандай буюртма бириктирилмаган.*", parse_mode="Markdown")
        return
    
    holat_map = {
        'yangi': '🆕 Янги',
        'qabul_qilingan': '🟡 Қабул қилинган',
        'ishda': '🔵 Иш жараёнида',
        'yakunlangan': '✅ Якунланган',
        'bekor_qilingan': '❌ Бекор қилинган',
        'rad_etilgan': '🚫 Рад этилган'
    }
    
    txt = "📋 *Менинг буюртмаларим:*\n\n"
    for o in orders[:10]:
        txt += f"🆔 #{o[0]} | {holat_map.get(o[6], o[6])}\n"
        txt += f"📝 {o[4][:40]}...\n\n"
    
    if len(orders) > 10:
        txt += f"\n... ва яна {len(orders)-10} та"
    
    # Ҳолатни ўзгартириш учун тугмачалар
    if orders:
        first_order = orders[0]
        if first_order[6] in ['qabul_qilingan', 'ishda']:
            txt += "\n\n⬇️ Ҳолатни ўзгартириш:"
            await message.answer(txt, reply_markup=buyurtma_holat_inline(first_order[0]), parse_mode="Markdown")
        else:
            await message.answer(txt, parse_mode="Markdown")

# ============ 18. УСТА: БУЮРТМА ҚАБУЛ ҚИЛИШ ============
@dp.message_handler(lambda msg: msg.text.startswith("/accept_"))
async def usta_accept_order(message: types.Message):
    usta = db.usta_get(tg_id=message.from_user.id)
    if not usta:
        await message.answer("⚠️ *Сиз уста эмассиз.*", parse_mode="Markdown")
        return
    
    buyurtma_id = int(message.text.split("_")[1])
    order = db.buyurtma_get(buyurtma_id)
    if not order:
        await message.answer("⚠️ *Буюртма топилмади.*", parse_mode="Markdown")
        return
    
    if order[6] != 'yangi':
        await message.answer("⚠️ *Буюртма аллақачон қабул қилинган.*", parse_mode="Markdown")
        return
    
    db.buyurtma_update_holat(buyurtma_id, 'qabul_qilingan', usta[0])
    await message.answer(f"✅ *Буюртма #{buyurtma_id} қабул қилинди!*", parse_mode="Markdown")

# ============ 19. ДИСПЕТЧЕР: БУЮРТМАЛАР ============
@dp.message_handler(lambda msg: msg.text in [
    "📋 Барча буюртмалар", "🆕 Янги буюртмалар", 
    "🟡 Қабул қилинганлар", "🔵 Ишдаги буюртмалар", 
    "✅ Якунланганлар", "❌ Бекор қилинганлар", "🚫 Рад этилганлар"
])
async def dispatcher_orders(message: types.Message):
    if message.from_user.id not in DISPATCHER_IDS and message.from_user.id not in ADMIN_IDS:
        return
    
    holat_map = {
        "📋 Барча буюртмалар": None,
        "🆕 Янги буюртмалар": "yangi",
        "🟡 Қабул қилинганлар": "qabul_qilingan",
        "🔵 Ишдаги буюртмалар": "ishda",
        "✅ Якунланганлар": "yakunlangan",
        "❌ Бекор қилинганлар": "bekor_qilingan",
        "🚫 Рад этилганлар": "rad_etilgan"
    }
    
    orders = db.buyurtma_all(holat_map.get(message.text))
    if not orders:
        await message.answer("📭 *Буюртмалар топилмади.*", parse_mode="Markdown")
        return
    
    holat_emoji = {
        'yangi': '🆕',
        'qabul_qilingan': '🟡',
        'ishda': '🔵',
        'yakunlangan': '✅',
        'bekor_qilingan': '❌',
        'rad_etilgan': '🚫'
    }
    
    txt = f"📋 *{message.text}:*\n\n"
    for o in orders[:10]:
        mijoz = db.mijoz_get(o[1])
        mijoz_ism = mijoz[2] if mijoz else "Noma'lum"
        txt += f"{holat_emoji.get(o[6], '📌')} #{o[0]} | {mijoz_ism}\n"
        txt += f"📝 {o[4][:30]}...\n\n"
    
    if len(orders) > 10:
        txt += f"\n... ва яна {len(orders)-10} та"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 20. ДИСПЕТЧЕР: СТАТИСТИКА ============
@dp.message_handler(lambda msg: msg.text == "📊 Статистика")
async def dispatcher_statistika(message: types.Message):
    if message.from_user.id not in DISPATCHER_IDS and message.from_user.id not in ADMIN_IDS:
        return
    
    stats = db.buyurtma_statistika()
    mijozlar = db.mijoz_all()
    ustalar = db.usta_barcha()
    
    if stats:
        txt = (
            f"📊 *Умумий статистика:*\n\n"
            f"👤 Мижозлар: {len(mijozlar)} та\n"
            f"👨‍🔧 Усталар: {len(ustalar)} та\n"
            f"📋 Жами буюртма: {stats[0]} та\n"
            f"🆕 Янги: {stats[1]} та\n"
            f"🟡 Қабул қилинган: {stats[2]} та\n"
            f"🔵 Ишда: {stats[3]} та\n"
            f"✅ Якунланган: {stats[4]} та\n"
            f"❌ Бекор қилинган: {stats[5]} та\n"
            f"🚫 Рад этилган: {stats[6]} та"
        )
        await message.answer(txt, parse_mode="Markdown")

# ============ 21. ДИСПЕТЧЕР: УСТАЛАР СТАТИСТИКАСИ ============
@dp.message_handler(lambda msg: msg.text == "👨‍🔧 Усталар статистикаси")
async def dispatcher_usta_statistika(message: types.Message):
    if message.from_user.id not in DISPATCHER_IDS and message.from_user.id not in ADMIN_IDS:
        return
    
    ustalar = db.usta_barcha()
    if not ustalar:
        await message.answer("📭 *Усталар мавжуд эмас.*", parse_mode="Markdown")
        return
    
    txt = "👨‍🔧 *Усталар статистикаси:*\n\n"
    for u in ustalar[:10]:
        holat = "🟢" if u[6] else "🔴"
        txt += f"{holat} {u[1]}\n🔧 {u[2]} | ⭐{u[4]:.1f} | 🏆{u[5]} та\n\n"
    
    if len(ustalar) > 10:
        txt += f"\n... ва яна {len(ustalar)-10} та уста"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 22. ДИСПЕТЧЕР: ОЙЛИК ҲИСОБОТ ============
@dp.message_handler(lambda msg: msg.text == "📈 Ойлик ҳисобот")
async def dispatcher_oylik_hisobot(message: types.Message):
    if message.from_user.id not in DISPATCHER_IDS and message.from_user.id not in ADMIN_IDS:
        return
    
    stats = db.buyurtma_oylik_statistika()
    if not stats:
        await message.answer("📭 *Маълумотлар мавжуд эмас.*", parse_mode="Markdown")
        return
    
    txt = "📈 *Ойлик ҳисобот:*\n\n"
    for s in stats:
        oy = s[0].strftime('%B %Y') if s[0] else "Noma'lum"
        txt += f"📅 {oy}: {s[1]} та буюртма, ✅ {s[2]} та якунланган\n"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 23. АДМИН: УСТА ҚЎШИШ ============
@dp.message_handler(lambda msg: msg.text == "👨‍🔧 Уста қўшиш" and msg.from_user.id in ADMIN_IDS)
async def admin_add_usta_start(message: types.Message, state: FSMContext):
    await message.answer("👨‍🔧 *Уста маълумотлари*\n\n📛 Исми:", parse_mode="Markdown")
    await AdminState.usta_qoshish_ism.set()

@dp.message_handler(state=AdminState.usta_qoshish_ism)
async def admin_add_usta_ism(message: types.Message, state: FSMContext):
    await state.update_data(ism=message.text)
    await message.answer("🔧 *Мутахассислиги:*", parse_mode="Markdown")
    await AdminState.usta_qoshish_mutaxassislik.set()

@dp.message_handler(state=AdminState.usta_qoshish_mutaxassislik)
async def admin_add_usta_mutaxassislik(message: types.Message, state: FSMContext):
    await state.update_data(mutaxassislik=message.text)
    await message.answer("📱 *Телефон рақами:*", parse_mode="Markdown")
    await AdminState.usta_qoshish_telefon.set()

@dp.message_handler(state=AdminState.usta_qoshish_telefon)
async def admin_add_usta_telefon(message: types.Message, state: FSMContext):
    await state.update_data(telefon=message.text)
    await message.answer("🆔 *Telegram IDси:*", parse_mode="Markdown")
    await AdminState.usta_qoshish_tg.set()

@dp.message_handler(state=AdminState.usta_qoshish_tg)
async def admin_add_usta_tg(message: types.Message, state: FSMContext):
    try:
        tg_id = int(message.text)
        data = await state.get_data()
        db.usta_qoshish(tg_id, data['ism'], data['mutaxassislik'], data['telefon'])
        await message.answer(
            f"✅ *Уста қўшилди!*\n\n"
            f"📛 {data['ism']}\n"
            f"🔧 {data['mutaxassislik']}\n"
            f"📱 {data['telefon']}\n"
            f"🆔 {tg_id}",
            parse_mode="Markdown"
        )
        await state.finish()
    except:
        await message.answer("⚠️ *Тўғри ID киритинг!*", parse_mode="Markdown")

# ============ 24. АДМИН: УСТА ЎЧИРИШ ============
@dp.message_handler(lambda msg: msg.text == "👨‍🔧 Уста ўчириш" and msg.from_user.id in ADMIN_IDS)
async def admin_delete_usta_start(message: types.Message):
    ustalar = db.usta_barcha()
    if not ustalar:
        await message.answer("📭 *Усталар мавжуд эмас.*", parse_mode="Markdown")
        return
    
    kb = InlineKeyboardMarkup(row_width=1)
    for u in ustalar:
        kb.add(InlineKeyboardButton(
            f"❌ {u[1]} - {u[2]} (ID:{u[1]})",
            callback_data=f"del_usta_{u[1]}"
        ))
    await message.answer("👨‍🔧 *Ўчириш учун устани танланг:*", reply_markup=kb, parse_mode="Markdown")

@dp.callback_query_handler(lambda c: c.data.startswith("del_usta_"))
async def admin_delete_usta(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[2])
    db.usta_ochirish(tg_id)
    await callback.message.edit_text(f"✅ *Уста ўчирилди!*", parse_mode="Markdown")

# ============ 25. АДМИН: ДИСПЕТЧЕР ҚЎШИШ ============
@dp.message_handler(lambda msg: msg.text == "👤 Диспетчер қўшиш" and msg.from_user.id in ADMIN_IDS)
async def admin_add_dispatcher_start(message: types.Message, state: FSMContext):
    await message.answer("🆔 *Диспетчернинг Telegram IDсини киритинг:*", parse_mode="Markdown")
    await AdminState.dispetcher_qoshish.set()

@dp.message_handler(state=AdminState.dispetcher_qoshish)
async def admin_add_dispatcher(message: types.Message, state: FSMContext):
    try:
        disp_id = int(message.text)
        if disp_id not in DISPATCHER_IDS:
            DISPATCHER_IDS.append(disp_id)
        await message.answer(f"✅ *Диспетчер қўшилди!* (ID: {disp_id})", parse_mode="Markdown")
        await state.finish()
    except:
        await message.answer("⚠️ *Тўғри ID киритинг!*", parse_mode="Markdown")

# ============ 26. АДМИН: ДИСПЕТЧЕР ЎЧИРИШ ============
@dp.message_handler(lambda msg: msg.text == "👤 Диспетчер ўчириш" and msg.from_user.id in ADMIN_IDS)
async def admin_delete_dispatcher(message: types.Message):
    if not DISPATCHER_IDS:
        await message.answer("📭 *Диспетчерлар мавжуд эмас.*", parse_mode="Markdown")
        return
    
    kb = InlineKeyboardMarkup(row_width=1)
    for d in DISPATCHER_IDS:
        try:
            user = await bot.get_chat(d)
            name = user.full_name
        except:
            name = f"ID:{d}"
        kb.add(InlineKeyboardButton(f"❌ {name}", callback_data=f"del_disp_{d}"))
    await message.answer("👤 *Ўчириш учун диспетчерни танланг:*", reply_markup=kb, parse_mode="Markdown")

@dp.callback_query_handler(lambda c: c.data.startswith("del_disp_"))
async def admin_delete_dispatcher_callback(callback: types.CallbackQuery):
    disp_id = int(callback.data.split("_")[2])
    if disp_id in DISPATCHER_IDS:
        DISPATCHER_IDS.remove(disp_id)
    await callback.message.edit_text(f"✅ *Диспетчер ўчирилди!* (ID: {disp_id})", parse_mode="Markdown")

# ============ 27. АДМИН: УМУМИЙ СТАТИСТИКА ============
@dp.message_handler(lambda msg: msg.text == "📊 Умумий статистика" and msg.from_user.id in ADMIN_IDS)
async def admin_statistika(message: types.Message):
    stats = db.buyurtma_statistika()
    mijozlar = db.mijoz_all()
    ustalar = db.usta_barcha()
    
    # Тўловлар статистикаси
    db.cursor.execute("SELECT COUNT(*), SUM(summa) FROM tolovlar WHERE holat='тўланди'")
    tolov_stats = db.cursor.fetchone()
    
    txt = (
        f"👑 *Админ статистикаси:*\n\n"
        f"👤 Мижозлар: {len(mijozlar)} та\n"
        f"👨‍🔧 Усталар: {len(ustalar)} та\n"
        f"📋 Жами буюртма: {stats[0]} та\n"
        f"🆕 Янги: {stats[1]} та\n"
        f"🟡 Қабул қилинган: {stats[2]} та\n"
        f"🔵 Ишда: {stats[3]} та\n"
        f"✅ Якунланган: {stats[4]} та\n"
        f"❌ Бекор қилинган: {stats[5]} та\n"
        f"🚫 Рад этилган: {stats[6]} та\n"
        f"💳 Тўловлар: {tolov_stats[0] or 0} та\n"
        f"💰 Жами тўлов: {tolov_stats[1] or 0:,.0f} сўм"
    )
    await message.answer(txt, parse_mode="Markdown")

# ============ 28. АДМИН: КУНЛИК ҲИСОБОТ ============
@dp.message_handler(lambda msg: msg.text == "📈 Кунлик ҳисобот" and msg.from_user.id in ADMIN_IDS)
async def admin_kunlik_hisobot(message: types.Message):
    today = datetime.now().date()
    orders = db.buyurtma_all()
    
    bugungi = [o for o in orders if o[8].date() == today]
    txt = (
        f"📈 *Кунлик ҳисобот*\n\n"
        f"📅 {today.strftime('%d.%m.%Y')}\n"
        f"📋 Жами буюртма: {len(bugungi)} та\n"
    )
    
    holatlar = {}
    for o in bugungi:
        holatlar[o[6]] = holatlar.get(o[6], 0) + 1
    
    for h, c in holatlar.items():
        emoji = {'yangi':'🆕','qabul_qilingan':'🟡','ishda':'🔵','yakunlangan':'✅','bekor_qilingan':'❌','rad_etilgan':'🚫'}
        txt += f"{emoji.get(h, '📌')} {h}: {c} та\n"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 29. АДМИН: ҲАФТАЛИК ҲИСОБОТ ============
@dp.message_handler(lambda msg: msg.text == "📈 Ҳафталик ҳисобот" and msg.from_user.id in ADMIN_IDS)
async def admin_haftalik_hisobot(message: types.Message):
    hafta = datetime.now().date() - timedelta(days=7)
    orders = db.buyurtma_all()
    
    haftalik = [o for o in orders if o[8].date() >= hafta]
    
    txt = (
        f"📈 *Ҳафталик ҳисобот*\n\n"
        f"📅 {hafta.strftime('%d.%m')} - {datetime.now().strftime('%d.%m.%Y')}\n"
        f"📋 Жами буюртма: {len(haftalik)} та\n"
    )
    
    holatlar = {}
    for o in haftalik:
        holatlar[o[6]] = holatlar.get(o[6], 0) + 1
    
    for h, c in holatlar.items():
        emoji = {'yangi':'🆕','qabul_qilingan':'🟡','ishda':'🔵','yakunlangan':'✅','bekor_qilingan':'❌','rad_etilgan':'🚫'}
        txt += f"{emoji.get(h, '📌')} {h}: {c} та\n"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 30. АДМИН: ОЙЛИК ҲИСОБОТ ============
@dp.message_handler(lambda msg: msg.text == "📈 Ойлик ҳисобот" and msg.from_user.id in ADMIN_IDS)
async def admin_oylik_hisobot(message: types.Message):
    oy = datetime.now().date() - timedelta(days=30)
    orders = db.buyurtma_all()
    
    oylik = [o for o in orders if o[8].date() >= oy]
    
    txt = (
        f"📈 *Ойлик ҳисобот*\n\n"
        f"📅 {oy.strftime('%d.%m')} - {datetime.now().strftime('%d.%m.%Y')}\n"
        f"📋 Жами буюртма: {len(oylik)} та\n"
    )
    
    holatlar = {}
    for o in oylik:
        holatlar[o[6]] = holatlar.get(o[6], 0) + 1
    
    for h, c in holatlar.items():
        emoji = {'yangi':'🆕','qabul_qilingan':'🟡','ishda':'🔵','yakunlangan':'✅','bekor_qilingan':'❌','rad_etilgan':'🚫'}
        txt += f"{emoji.get(h, '📌')} {h}: {c} та\n"
    
    await message.answer(txt, parse_mode="Markdown")

# ============ 31. АДМИН: ХАБАР ЮБОРИШ ============
@dp.message_handler(lambda msg: msg.text in ["📢 Барчага хабар", "📢 Усталарга хабар", "📢 Мижозларга хабар"] and msg.from_user.id in ADMIN_IDS)
async def admin_xabar_start(message: types.Message, state: FSMContext):
    kimgatur = {
        "📢 Барчага хабар": "all",
        "📢 Усталарга хабар": "ustalar",
        "📢 Мижозларга хабар": "mijozlar"
    }
    await state.update_data(kimgatur=kimgatur[message.text])
    await message.answer("📝 *Хабар матнини киритинг:*", parse_mode="Markdown")
    await AdminState.xabar_yuborish.set()

@dp.message_handler(state=AdminState.xabar_yuborish)
async def admin_xabar_send(message: types.Message, state: FSMContext):
    data = await state.get_data()
    matn = message.text
    kimgatur = data['kimgatur']
    
    db.xabar_save(message.from_user.id, kimgatur, matn)
    
    count = 0
    if kimgatur == "all":
        for m in db.mijoz_all():
            try:
                await bot.send_message(m[1], f"📢 *Хабар:*\n\n{matn}", parse_mode="Markdown")
                count += 1
            except:
                pass
        for u in db.usta_barcha():
            try:
                await bot.send_message(u[1], f"📢 *Хабар:*\n\n{matn}", parse_mode="Markdown")
                count += 1
            except:
                pass
    elif kimgatur == "ustalar":
        for u in db.usta_barcha():
            try:
                await bot.send_message(u[1], f"📢 *Усталарга хабар:*\n\n{matn}", parse_mode="Markdown")
                count += 1
            except:
                pass
    elif kimgatur == "mijozlar":
        for m in db.mijoz_all():
            try:
                await bot.send_message(m[1], f"📢 *Мижозларга хабар:*\n\n{matn}", parse_mode="Markdown")
                count += 1
            except:
                pass
    
    await message.answer(f"✅ *Хабар юборилди!* ({count} та фойдаланувчига)", parse_mode="Markdown")
    await state.finish()

# ============ 32. АДМИН: EXCEL ҲИСОБОТ ============
@dp.message_handler(lambda msg: msg.text == "📥 Excel ҳисобот" and msg.from_user.id in ADMIN_IDS)
async def admin_excel_report(message: types.Message):
    await message.answer("📥 *Excel ҳисобот тайёрланмоқда...*", parse_mode="Markdown")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # 1. Буюртмалар варағи
        ws1 = wb.active
        ws1.title = "Буюртмалар"
        
        headers = ["ID", "Мижоз", "Телефон", "Уста", "Матн", "Манзил", "Нарх", "Ҳолат", "Вақт"]
        ws1.append(headers)
        
        # Сарлавҳа стили
        for col in range(1, len(headers)+1):
            ws1.cell(row=1, column=col).font = Font(bold=True)
            ws1.cell(row=1, column=col).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        orders = db.buyurtma_all()
        for o in orders:
            mijoz = db.mijoz_get(o[1])
            usta = db.usta_get(usta_id=o[2]) if o[2] else None
            ws1.append([
                o[0],
                mijoz[2] if mijoz else "Noma'lum",
                mijoz[3] if mijoz else "",
                usta[1] if usta else "Ўтмаган",
                o[4],
                o[5] or "",
                float(o[6]) if o[6] else 0,
                o[7],
                o[8].strftime("%d.%m.%Y %H:%M") if o[8] else ""
            ])
        
        # 2. Усталар варағи
        ws2 = wb.create_sheet("Усталар")
        ws2.append(["ID", "Исм", "Мутахассислик", "Телефон", "Рейтинг", "Бажарган", "Ҳолат"])
        for u in db.usta_barcha():
            ws2.append([
                u[0], u[1], u[2], u[3], u[4], u[5], "Онлайн" if u[6] else "Банд"
            ])
        
        # 3. Статистика варағи
        ws3 = wb.create_sheet("Статистика")
        stats = db.buyurtma_statistika()
        ws3.append(["Кўрсаткич", "Қиймат"])
        ws3.append(["Жами буюртма", stats[0]])
        ws3.append(["Янги", stats[1]])
        ws3.append(["Қабул қилинган", stats[2]])
        ws3.append(["Ишда", stats[3]])
        ws3.append(["Якунланган", stats[4]])
        ws3.append(["Бекор қилинган", stats[5]])
        ws3.append(["Рад этилган", stats[6]])
        
        filename = f"hisobot_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        wb.save(filename)
        
        await message.answer_document(
            open(filename, 'rb'),
            caption=f"📊 *Ҳисобот {datetime.now().strftime('%d.%m.%Y')}*\n\n"
                   f"📋 Жами: {stats[0]} та буюртма\n"
                   f"✅ Якунланган: {stats[4]} та",
            parse_mode="Markdown"
        )
        os.remove(filename)
    except ImportError:
        await message.answer("⚠️ *Excel учун openpyxl ўрнатинг: pip install openpyxl*", parse_mode="Markdown")
    except Exception as e:
        await message.answer(f"⚠️ *Хатолик: {e}*", parse_mode="Markdown")

# ============ 33. CALLBACK: УСТА ТАНЛАШ ============
@dp.callback_query_handler(lambda c: c.data.startswith("assign_"))
async def assign_usta(callback: types.CallbackQuery):
    _, usta_id, buyurtma_id = callback.data.split("_")
    usta_id = int(usta_id)
    buyurtma_id = int(buyurtma_id)
    
    db.buyurtma_update_holat(buyurtma_id, 'qabul_qilingan', usta_id)
    
    await callback.message.edit_text(
        f"✅ *Буюртма #{buyurtma_id} устага бириктирилди!*",
        parse_mode="Markdown"
    )
    
    usta = db.usta_get(usta_id=usta_id)
    if usta:
        try:
            await bot.send_message(
                usta[1],
                f"🟡 *Янги буюртма бириктирилди!*\n\n"
                f"🆔 #{buyurtma_id}\n\n"
                f"📋 Буюртмани кўриш учун /order_{buyurtma_id}",
                parse_mode="Markdown"
            )
        except:
            pass

# ============ 34. CALLBACK: ҲОЛАТ ЎЗГАРТИРИШ ============
@dp.callback_query_handler(lambda c: c.data.startswith(("qabul_", "rad_", "ishla_", "yakun_", "bekor_")))
async def holat_change(callback: types.CallbackQuery):
    action, buyurtma_id = callback.data.split("_")
    buyurtma_id = int(buyurtma_id)
    
    holat_map = {
        "qabul": "qabul_qilingan",
        "rad": "rad_etilgan",
        "ishla": "ishda",
        "yakun": "yakunlangan",
        "bekor": "bekor_qilingan"
    }
    
    holat = holat_map.get(action)
    if not holat:
        return
    
    order = db.buyurtma_get(buyurtma_id)
    if not order:
        await callback.message.edit_text("⚠️ *Буюртма топилмади.*", parse_mode="Markdown")
        return
    
    usta_id = order[2]
    db.buyurtma_update_holat(buyurtma_id, holat, usta_id)
    
    emoji = {
        "qabul_qilingan": "🟡",
        "rad_etilgan": "🚫",
        "ishda": "🔵",
        "yakunlangan": "✅",
        "bekor_qilingan": "❌"
    }
    
    await callback.message.edit_text(
        f"{emoji.get(holat, '📌')} *Буюртма #{buyurtma_id} ҳолати: {holat}*",
        parse_mode="Markdown"
    )
    
    try:
        await bot.send_message(
            order[1],
            f"📌 *Буюртма #{buyurtma_id} ҳолати:*\n{emoji.get(holat, '📌')} {holat}",
            parse_mode="Markdown"
        )
    except:
        pass

# ============ 35. CALLBACK: БОШҚА УСТАГА ============
@dp.callback_query_handler(lambda c: c.data.startswith("change_"))
async def change_usta_start(callback: types.CallbackQuery, state: FSMContext):
    buyurtma_id = int(callback.data.split("_")[1])
    await state.update_data(buyurtma_id=buyurtma_id)
    
    ustalar = db.usta_top()
    if not ustalar:
        await callback.message.edit_text("📭 *Усталар мавжуд эмас.*", parse_mode="Markdown")
        return
    
    await callback.message.edit_text(
        "🔄 *Бошқа уста танланг:*",
        reply_markup=usta_tanlash_inline(ustalar, buyurtma_id),
        parse_mode="Markdown"
    )
    await ChangeUstaState.usta_tanlash.set()

@dp.callback_query_handler(lambda c: c.data.startswith("assign_"), state=ChangeUstaState.usta_tanlash)
async def change_usta_assign(callback: types.CallbackQuery, state: FSMContext):
    _, usta_id, buyurtma_id = callback.data.split("_")
    usta_id = int(usta_id)
    buyurtma_id = int(buyurtma_id)
    
    db.buyurtma_usta_change(buyurtma_id, usta_id)
    await callback.message.edit_text(
        f"🔄 *Буюртма #{buyurtma_id} бошқа устага берилди!*",
        parse_mode="Markdown"
    )
    await state.finish()

# ============ 36. БУЮРТМАНИ КЎРИШ (/order_123) ============
@dp.message_handler(lambda msg: msg.text.startswith("/order_"))
async def view_order(message: types.Message):
    buyurtma_id = int(message.text.split("_")[1])
    order = db.buyurtma_get(buyurtma_id)
    if not order:
        await message.answer("⚠️ *Буюртма топилмади.*", parse_mode="Markdown")
        return
    
    holat_map = {
        'yangi': '🆕 Янги',
        'qabul_qilingan': '🟡 Қабул қилинган',
        'ishda': '🔵 Иш жараёнида',
        'yakunlangan': '✅ Якунланган',
        'bekor_qilingan': '❌ Бекор қилинган',
        'rad_etilgan': '🚫 Рад этилган'
    }
    
    mijoz = db.mijoz_get(order[1])
    usta = db.usta_get(usta_id=order[2]) if order[2] else None
    
    txt = (
        f"🆔 *Буюртма №{buyurtma_id}*\n\n"
        f"👤 Мижоз: {mijoz[2] if mijoz else 'Noma'lum'}\n"
        f"📱 Телефон: {mijoz[3] if mijoz else 'Noma'lum'}\n"
        f"📍 Манзил: {order[5] or 'Кўрсатилмаган'}\n"
        f"📝 Матн: {order[4]}\n"
        f"💰 Нарх: {order[6] or 0} сўм\n"
        f"📌 Ҳолат: {holat_map.get(order[7], order[7])}\n"
        f"👨‍🔧 Уста: {usta[1] if usta else 'Танланмаган'}\n"
        f"📅 Вақт: {order[8].strftime('%d.%m.%Y %H:%M') if order[8] else ''}"
    )
    
    # Ҳолатни ўзгартириш учун тугмачалар
    if order[7] in ['qabul_qilingan', 'ishda']:
        await message.answer(txt, reply_markup=buyurtma_holat_inline(buyurtma_id), parse_mode="Markdown")
    else:
        await message.answer(txt, parse_mode="Markdown")

# ============ 37. ЭСЛАТМАЛАРНИ ТЕКШИРИШ ============
async def check_eslatmalar():
    while True:
        try:
            now = datetime.now()
            eslatmalar = db.eslatma_get_vaqt(now)
            
            for e in eslatmalar:
                try:
                    await bot.send_message(
                        e[1],
                        f"🔔 *Эслатма!*\n\n📝 {e[2]}",
                        parse_mode="Markdown"
                    )
                    db.eslatma_ochirish(e[0])
                except:
                    pass
            
            await asyncio.sleep(30)
        except:
            await asyncio.sleep(60)

# ============ 38. ТЕСТ УСТАЛАР ============
async def on_startup():
    print("🚀 Бот ишга тушди!")
    if not db.usta_barcha():
        db.usta_qoshish(0, "Бобур", "Компьютер таъмирлаш", "+998901234567")
        db.usta_qoshish(0, "Жамшид", "Электр монтаж", "+998901234568")
        db.usta_qoshish(0, "Дилшод", "Сантехник", "+998901234569")
        print("✅ Тест усталар қўшилди!")
    asyncio.create_task(check_eslatmalar())

# ============ 39. ИШГА ТУШИРИШ ============
if __name__ == "__main__":
    from aiogram import executor
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)
